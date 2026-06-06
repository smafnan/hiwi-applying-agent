import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fill colors for relevance scores
FILL_HIGH = PatternFill("solid", fgColor="C6EFCE")    # green — score ≥ 60
FILL_MID  = PatternFill("solid", fgColor="FFEB9C")    # yellow — score 30–59
FILL_LOW  = PatternFill("solid", fgColor="F2F2F2")    # gray — score < 30
FILL_HDR  = PatternFill("solid", fgColor="2F4F8F")    # dark blue — header

FONT_HDR  = Font(bold=True, color="FFFFFF", size=11)
FONT_BODY = Font(size=10)
FONT_LINK = Font(size=10, color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# v2 upgraded column set
COLUMNS = [
    ("Name",              30),
    ("Department",        25),
    ("Email",             32),
    ("Profile URL",       30),
    ("Source Courses",    40),
    ("Score /100",        12),
    ("Breakdown",         30),
    ("Language",          12),
    ("Grant Indicators",  35),
    ("Research area",     50),
    ("CV connection",     55),
    ("Email draft",       80),
    ("Review passed",     15),
    ("Status",            14),
    ("Follow-up Day 7",   70),
    ("Follow-up Day 18",  70),
    ("Notes",             30),
]


def write_sheet(professors: list) -> None:
    Path("output").mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach tracker"

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HDR
        cell.font = FONT_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Sort: highest score first, unscored last
    def sort_key(p):
        return -(p.get("relevance_score") or 0)

    sorted_profs = sorted(professors, key=sort_key)

    for row_idx, prof in enumerate(sorted_profs, start=2):
        score = prof.get("relevance_score") or 0
        if score >= 60:
            row_fill = FILL_HIGH
        elif score >= 30:
            row_fill = FILL_MID
        else:
            row_fill = FILL_LOW

        courses_str = "; ".join(prof.get("source_courses", []))
        projects_str = "; ".join(prof.get("current_projects", []))
        research_str = prof.get("research_summary") or ""
        if projects_str:
            research_str = research_str + "\n\nProjects: " + projects_str

        breakdown = prof.get("relevance_breakdown", {})
        breakdown_str = (
            f"course:{breakdown.get('course_taken',0)} "
            f"AI:{breakdown.get('ai_overlap',0)} "
            f"grant:{breakdown.get('active_grant',0)} "
            f"posting:{breakdown.get('open_position',0)} "
            f"pubs:{breakdown.get('has_publications',0)}"
        )
        grant_str = "; ".join(
            f"{g.get('funder','')} {g.get('project','')}"
            for g in prof.get("grant_indicators", [])
        )
        review_passed = "✓" if prof.get("review_result", {}).get("passed") else "✗"

        values = [
            prof.get("name", ""),
            prof.get("department", ""),
            prof.get("email", ""),
            prof.get("profile_url", ""),
            courses_str,
            prof.get("relevance_score") or "",
            breakdown_str,
            prof.get("email_language", "english"),
            grant_str,
            research_str,
            prof.get("cv_connection", ""),
            prof.get("email_draft", ""),
            review_passed,
            prof.get("status", "draft"),
            prof.get("followup_day7", ""),
            prof.get("followup_day18", ""),
            "",  # Notes — blank for user to fill
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="left"
            )
            cell.font = FONT_BODY

        # Make email a mailto hyperlink (column 3)
        email = prof.get("email")
        if email:
            email_cell = ws.cell(row=row_idx, column=3)
            email_cell.hyperlink = f"mailto:{email}"
            email_cell.font = FONT_LINK

        # Make profile URL a clickable link (column 4)
        url = prof.get("profile_url")
        if url:
            url_cell = ws.cell(row=row_idx, column=4)
            url_cell.hyperlink = url
            url_cell.font = FONT_LINK

        # Row height based on email draft length
        # (email_draft is initialised to None, so guard with `or ""` — len(None) crashes)
        draft_len = len(prof.get("email_draft") or "")
        ws.row_dimensions[row_idx].height = max(60, min(200, draft_len // 3))

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a summary sheet
    ws_summary = wb.create_sheet("Summary")
    score3 = sum(1 for p in professors if (p.get("relevance_score") or 0) >= 60)
    score2 = sum(1 for p in professors if 30 <= (p.get("relevance_score") or 0) < 60)
    score1 = sum(1 for p in professors if (p.get("relevance_score") or 0) < 30)
    with_email = sum(1 for p in professors if p.get("email"))
    with_draft = sum(1 for p in professors if p.get("email_draft"))
    de = sum(1 for p in professors if p.get("email_language") == "german" and p.get("email_draft"))
    en = sum(1 for p in professors if p.get("email_language") == "english" and p.get("email_draft"))
    flagged = sum(1 for p in professors if p.get("status") == "needs_manual_review")

    summary_rows = [
        ("Metric", "Count"),
        ("Total professors found", len(professors)),
        ("Professors with email", with_email),
        ("Emails drafted", with_draft),
        ("German emails", de),
        ("English emails", en),
        ("High fit (≥60)", score3),
        ("Medium fit (30–59)", score2),
        ("Low fit (<30)", score1),
        ("Flagged for manual review", flagged),
    ]
    for r_idx, (label, val) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=r_idx, column=1, value=label)
        ws_summary.cell(row=r_idx, column=2, value=val)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 12

    output_path = "output/outreach_tracker.xlsx"
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    profs_path = Path("data/professors.json")
    if not profs_path.exists():
        print("ERROR: professors.json not found.")
        raise SystemExit(1)

    with open(profs_path) as f:
        professors = json.load(f)

    print(f"Writing sheet for {len(professors)} professors...")
    path = write_sheet(professors)
    print(f"\n✓ Output saved to {path}")
    print("Open outreach_tracker.xlsx — sort by Score, review top emails first.")
